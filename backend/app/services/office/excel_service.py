import logging

logger = logging.getLogger(__name__)


class ExcelService:

    async def export(self, file_path: str, json_data: dict) -> None:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

        sheets = json_data.get("sheets", [])
        if not sheets:
            inner = json_data.get("content", {})
            sheets = inner.get("sheets", []) if inner else []

        if not sheets and "blocks" in json_data:
            table_text = ""
            for block in json_data["blocks"]:
                if block.get("type") == "table":
                    table_text = block.get("text", "")
                    break
            if table_text:
                ws = wb.create_sheet(title="Sheet1")
                for row_idx, line in enumerate(table_text.split("\n"), 1):
                    if line.startswith("[Sheet:"):
                        continue
                    for col_idx, cell in enumerate(line.split(" | "), 1):
                        ws.cell(row=row_idx, column=col_idx, value=cell.strip())

        for sheet_data in sheets:
            ws = wb.create_sheet(title=sheet_data.get("name", "Sheet1"))
            cells = sheet_data.get("cells", {})
            for coord, info in cells.items():
                if info.get("type") == "null":
                    continue
                ws[coord] = info.get("value")

        wb.save(file_path)
