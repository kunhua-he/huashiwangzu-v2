"""Sandbox test for excel-engine module.

Tests the parsing, state management, and generation pipeline.
"""
import os
import sys
import json
import tempfile
import asyncio

# Add module root to path (for `from backend.xxx` imports)
_MODULE_ROOT = os.path.dirname(os.path.dirname(__file__))  # modules/excel-engine/
sys.path.insert(0, _MODULE_ROOT)

# Test constants
TEST_DATA = {
    'cells': {
        'A1': '姓名', 'B1': '年龄', 'C1': '城市',
        'A2': '张三', 'B2': '28', 'C2': '北京',
        'A3': '李四', 'B3': '32', 'C3': '上海',
    },
    'styles': {
        'A1': {'bold': True, 'fontName': '微软雅黑', 'fontSize': 12, 'fillColor': '#4472C4', 'color': '#FFFFFF'},
    },
    'merges': {},
    'col_widths': {'A': 100, 'B': 60},
    'row_heights': {},
    'total_rows': 40,
    'total_cols': 10,
}


def test_address_tool():
    """Test address utility functions"""
    from backend.tool.address import parse_address, rc_to_address, col_letter

    assert parse_address('A1') == {'r': 1, 'c': 0}
    assert parse_address('B8') == {'r': 8, 'c': 1}
    assert parse_address('AA2') == {'r': 2, 'c': 26}

    assert rc_to_address(1, 0) == 'A1'
    assert rc_to_address(5, 1) == 'B5'
    assert rc_to_address(3, 26) == 'AA3'

    assert col_letter(0) == 'A'
    assert col_letter(25) == 'Z'
    assert col_letter(26) == 'AA'

    print('  ✓ test_address_tool passed')


def test_formula():
    """Test formula calculation"""
    from backend.tool.formula import calculate_formula

    cells = {'A1': '10', 'A2': '20', 'A3': '30', 'B1': '5'}
    assert calculate_formula('=SUM(A1:A3)', cells) == '60.0'
    assert calculate_formula('=AVERAGE(A1:A3)', cells) == '20.0'
    assert calculate_formula('=COUNT(A1:A3)', cells) == '3'
    assert calculate_formula('=MAX(A1:A3)', cells) == '30.0'
    assert calculate_formula('=MIN(A1:A3)', cells) == '10.0'
    assert calculate_formula('=A1+B1', cells) == '15'
    assert calculate_formula('=A1*B1', cells) == '50'
    assert calculate_formula('=A2/B1', cells) == '4.0'
    assert calculate_formula('hello', cells) == 'hello'

    print('  ✓ test_formula passed')


def test_state_manager():
    """Test state manager operations"""
    from backend.state.manager import cell_set_text, cell_get_text, cell_get_style_ref, cell_set_style_val
    from backend.state.db_ops import empty_state

    state = empty_state()
    assert state['total_rows'] == 40
    assert state['total_cols'] == 10

    # Cell operations
    cell_set_text(state, 'A1', 'Hello')
    assert cell_get_text(state, 'A1') == 'Hello'

    cell_set_text(state, 'A1', '')
    assert cell_get_text(state, 'A1') == ''

    # Style operations
    cell_set_style_val(state, 'B1', 'bold', True)
    assert cell_get_style_ref(state, 'B1')['bold'] is True

    print('  ✓ test_state_manager passed')


def test_xlsx_roundtrip():
    """Test XLSX generation and re-parsing"""
    from backend.engine.xlsx_generator import generate_xlsx

    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)

    try:
        success = generate_xlsx(tmp_path, {'Sheet1': TEST_DATA})
        assert success, "XLSX generation failed"
        assert os.path.getsize(tmp_path) > 100, "Generated file too small"

        # Verify file is valid by re-opening with openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        ws = wb['Sheet1']
        assert ws is not None
        assert ws['A1'].value == '姓名'
        assert ws['A2'].value == '张三'
        wb.close()

        print('  ✓ test_xlsx_roundtrip passed')
    finally:
        os.unlink(tmp_path)


def test_csv_generation():
    """Test CSV generation"""
    from backend.engine.xlsx_generator import generate_csv

    fd, tmp_path = tempfile.mkstemp(suffix='.csv')
    os.close(fd)

    try:
        csv_content = generate_csv(tmp_path, TEST_DATA)
        assert '姓名' in csv_content
        assert '张三' in csv_content
        assert os.path.getsize(tmp_path) > 20

        print('  ✓ test_csv_generation passed')
    finally:
        os.unlink(tmp_path)


async def test_edit_operations():
    """Test edit operations"""
    from backend.table.edit import EditOperations
    from backend.state.manager import cell_get_text

    state = dict(TEST_DATA)
    result = await EditOperations._input(state, 'test_key', ['A1'], {'value': '测试值'})
    assert result['code'] == 0
    assert state['cells']['A1'] == '测试值'

    result = await EditOperations._clear(state, 'test_key', ['A1', 'B1'], {'type': 'all'})
    assert result['code'] == 0
    assert cell_get_text(state, 'A1') == ''
    assert 'A1' not in state['cells']

    print('  ✓ test_edit_operations passed')


async def test_style_operations():
    """Test style operations"""
    from backend.table.style_ops import StyleOperations
    from backend.state.manager import cell_get_style_ref

    state = dict(TEST_DATA)
    result = await StyleOperations._toggle_style(state, ['A1'], 'bold')
    assert result['code'] == 0
    # Toggle again should be off
    await StyleOperations._toggle_style(state, ['A1'], 'bold')
    assert state['styles'].get('A1', {}).get('bold') is False

    print('  ✓ test_style_operations passed')


async def test_row_col_operations():
    """Test row/column operations"""
    from backend.table.row_col import RowColOperations
    from backend.tool.address import rc_to_address

    state = dict(TEST_DATA)
    old_a2_value = state.get('cells', {}).get('A2', '')
    result = await RowColOperations._delete_row(state, ['A2'])
    assert result['code'] == 0
    # Row 2 deleted: old A2 value gone, cells above moved down
    assert state.get('cells', {}).get('A2', '') != old_a2_value
    assert state.get('total_rows', 0) == 39

    print('  ✓ test_row_col_operations passed')


async def test_clipboard():
    """Test clipboard operations"""
    from backend.table.clipboard import ClipboardOperations

    state = dict(TEST_DATA)
    state['_clipboard'] = {}
    state['_clipboard_range'] = []

    result = ClipboardOperations._copy(state, ['A1', 'B1'])
    assert result['code'] == 0
    assert 'A1' in state.get('_clipboard', {})
    assert 'B1' in state.get('_clipboard', {})

    print('  ✓ test_clipboard passed')


if __name__ == '__main__':
    print('\n=== Excel Engine Module Tests ===\n')
    test_address_tool()
    test_formula()
    test_state_manager()
    test_xlsx_roundtrip()
    test_csv_generation()
    asyncio.run(test_edit_operations())
    asyncio.run(test_style_operations())
    asyncio.run(test_row_col_operations())
    asyncio.run(test_clipboard())
    print('\n✓ All tests passed!\n')
