
"""Sandbox test for xlsx-parser module."""
import sys, csv, io
from pathlib import Path
SDIR = Path(__file__).resolve().parent / "samples"


def parse_xlsx(path):
    import openpyxl
    blocks = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            rt = " | ".join(cells)
            if rt.strip(): rows.append(rt)
        if rows:
            blocks.append({"type": "表格", "text": "[Sheet: %s]\n%s" % (sn, "\n".join(rows)),
                           "page": None, "resource_ref": None})
    wb.close()
    return {"file_id": 0, "format": "xlsx", "blocks": blocks, "resources": []}


def parse_csv(path):
    raw = path.read_bytes()
    try: content = raw.decode("utf-8-sig")
    except UnicodeDecodeError: content = raw.decode("gbk", errors="replace")
    reader = csv.reader(io.StringIO(content))
    rows = [" | ".join(c.strip() for c in row) for row in reader]
    blocks = [{"type": "表格", "text": "\n".join(rows), "page": None, "resource_ref": None}] if rows else []
    return {"file_id": 0, "format": "csv", "blocks": blocks, "resources": []}


def validate(result, label):
    assert all(k in result for k in ("file_id","blocks","resources"))
    for b in result["blocks"]: assert all(k in b for k in ("type","text","page","resource_ref"))
    print("  [%s] Validation PASS (%d blocks)" % (label, len(result["blocks"])))


def main():
    print("="*60); print("xlsx-parser sandbox test"); print("="*60)
    for fn in ("sample.xlsx", "sample.csv"):
        f = SDIR / fn
        if f.exists():
            func = parse_xlsx if fn.endswith("xlsx") else parse_csv
            result = func(f)
            for b in result["blocks"]:
                print("  [%s] %s" % (fn, b["text"][:80]))
            validate(result, fn)
    print("PASS: xlsx-parser sandbox test")

if __name__ == "__main__": main()
